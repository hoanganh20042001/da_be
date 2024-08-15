from typing import Any, List, Dict
from multidict import MultiDict
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Series, Reference
import pandas as pd
import logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
class Excel:
    def __init__(self):
        self.wb = Workbook()

    def writeToExcelFile(self, data: Any, file_name: str, title: str):
        work_sheet = self.wb.active
        work_sheet.title = title
        for row in range(1, 40):
            work_sheet.append(range(1000))
        self.save(filename=file_name)


    def readExcelFile(self, filename: str, data:str):
        self.wb = load_workbook(filename=filename)
        sheet_ranges = self.wb[data.get('index').value]

    def exportToTemplate(self, data: List[List[Any]], template_file: str, output_file: str):
        # Tải template file
        wb = load_workbook(template_file)
        work_sheet = wb.active
        
        # Xóa dữ liệu hiện tại (nếu có) và ghi dữ liệu mới
        for row in work_sheet.iter_rows(min_row=2, max_col=work_sheet.max_column, max_row=work_sheet.max_row):
            for cell in row:
                cell.value = None  # Xóa dữ liệu hiện tại
        logger.info(data)
        for row in data:
            work_sheet.append(row)  # Thêm dữ liệu vào template
        
        # Lưu file Excel mới
        wb.save(output_file)

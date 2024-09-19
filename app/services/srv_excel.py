from fastapi import APIRouter, File, UploadFile, HTTPException
import pandas as pd
from typing import List, Dict, Any
from openpyxl import load_workbook
from fastapi.security import HTTPBearer
from fastapi_sqlalchemy import db
from app.models import Patients, Units
from datetime import datetime
import os
import logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
from app.helpers.excel import Excel
class ExcelService:
    __instance = None

    def __init__(self) -> None:
        pass

    reusable_oauth2 = HTTPBearer(
        scheme_name='Authorization'
    )


    @staticmethod
    async def import_patients(file: UploadFile):
        logger.info(file)
        contents =await file.read()
        
        df = pd.read_excel(contents)
        logger.info(df)
        # if not {'name', 'email'}.issubset(df.columns):
     
        with db():
            for _, row in df.iterrows():
                logger.info(row[3])
                unit=db.session.query(Units).filter(Units.name == row[5]).first()
                date_str = str(row[3]) 
                patient = Patients(full_name=row[1], 
                                   identification=row[2], 
                                   phone_number=row[4],
                                   date_birth=datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S"),
                                   sex = 1 if row[6] == 'Nam' else 0,
                                   resident=row[7],
                                   home_town=row[8],
                                   medical_history=row[13],
                                   rank=row[9],
                                   weight=row[10],
                                   height=row[11],
                                   blood_group=row[12],
                                   email=row[14],
                                   position=row[15],
                                   unit_id=unit.id,
                                   )
                db.session.add(patient)
            db.session.commit()
        return file.filename

    @staticmethod
    async def export_data_to_excel(template_file: str, output_folder: str):
        logger.info(1)
        data = db.session.query(Patients.full_name, Patients.phone_number).all()
        
        # Chuyển đổi dữ liệu thành danh sách danh sách
        data_list = [list(row) for row in data]
        logger.info(data_list)
        # Tạo tên file đầu ra với thời gian hiện tại
        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(output_folder, f"{now}.xlsx")

        # Kiểm tra nếu thư mục có sẵn
        if not os.path.exists(output_folder):
            raise HTTPException(status_code=400, detail=f"Output folder '{output_folder}' does not exist")

        # Xuất dữ liệu vào template file
        wb = load_workbook(template_file)
        work_sheet = wb.active
        
        for row in work_sheet.iter_rows(min_row=2, max_col=work_sheet.max_column, max_row=work_sheet.max_row):
            for cell in row:
                cell.value = None  # Clear existing data
        
        # Assuming the header row is the first row
        header_row = [cell.value for cell in work_sheet[1]]
        
        # Append new data
        for row_data in data_list:
            work_sheet.append(row_data)
        
        # Create the full path for the output file
        output_file_path = os.path.join(output_folder, output_file)
        
        # Save the new Excel file
        wb.save(output_file_path)

        logger.info(f"Exported data to {output_file_path}")

        return output_file_path